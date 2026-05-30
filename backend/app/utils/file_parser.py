import csv
import io
from pathlib import Path

from openpyxl import load_workbook

from app.services.storage_service import StorageService
from app.utils.validators import normalize_column


def _cell_value(value):
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None
    return value


def _rows_from_csv(data: bytes | str) -> list[dict]:
    if isinstance(data, bytes):
        text = data.decode("utf-8-sig")
    else:
        text = data
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        return []
    columns = [normalize_column(c) for c in reader.fieldnames]
    rows = []
    for raw in reader:
        rows.append(
            {col: _cell_value(raw.get(orig)) for orig, col in zip(reader.fieldnames, columns)}
        )
    return rows


def _rows_from_xlsx(data: bytes) -> list[dict]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        return []
    columns = [normalize_column(str(c) if c is not None else "") for c in header]
    rows = []
    for values in rows_iter:
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in values):
            continue
        row = {}
        for col, val in zip(columns, values):
            row[col] = _cell_value(val)
        rows.append(row)
    wb.close()
    return rows


def read_tabular_bytes(data: bytes, suffix: str) -> list[dict]:
    suffix = suffix.lower()
    if suffix == ".csv":
        return _rows_from_csv(data)
    if suffix in (".xlsx", ".xls"):
        if suffix == ".xls":
            raise ValueError("Legacy .xls not supported; save as .xlsx or .csv")
        return _rows_from_xlsx(data)
    raise ValueError(f"Unsupported file type: {suffix}")


def read_tabular_file(file_path: str) -> list[dict]:
    if file_path.startswith("supabase://"):
        data = StorageService.download_bytes(file_path)
        suffix = Path(StorageService.parse_uri(file_path)[1]).suffix.lower()
        return read_tabular_bytes(data, suffix)
    path = Path(file_path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _rows_from_csv(path.read_text(encoding="utf-8-sig"))
    if suffix in (".xlsx", ".xls"):
        return read_tabular_bytes(path.read_bytes(), suffix)
    raise ValueError(f"Unsupported file type: {suffix}")


def read_upload_file(file_storage) -> list[dict]:
    original = file_storage.filename or "file.csv"
    suffix = Path(original).suffix.lower()
    data = file_storage.read()
    file_storage.stream.seek(0)
    return read_tabular_bytes(data, suffix)
